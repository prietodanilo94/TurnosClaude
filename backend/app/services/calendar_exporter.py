"""
Genera el Excel de vista calendario (turno por semana ISO) para enviar al jefe de sucursal.
"""
from __future__ import annotations

import calendar as _cal
from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from app.models.schemas import CalendarExportRequest

MONTH_NAMES_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
WEEKDAY_ABBR = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

_FILL_TITLE   = PatternFill("solid", fgColor="1F4E79")
_FILL_WEEK    = PatternFill("solid", fgColor="2E75B6")
_FILL_HEADER  = PatternFill("solid", fgColor="BDD7EE")
_FILL_WEEKEND = PatternFill("solid", fgColor="FCE4D6")
_FILL_ALT     = PatternFill("solid", fgColor="F5F5F5")
_FILL_NONE    = PatternFill("none")

_FONT_TITLE   = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
_FONT_WEEK    = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_FONT_HEADER  = Font(name="Calibri", bold=True, size=9)
_FONT_NAME    = Font(name="Calibri", bold=True, size=9)
_FONT_SHIFT   = Font(name="Calibri", size=9)
_FONT_LIBRE   = Font(name="Calibri", size=9, color="AAAAAA", italic=True)
_FONT_HOURS   = Font(name="Calibri", bold=True, size=9)

_THIN = Side(style="thin", color="CCCCCC")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _dias_iso_del_mes(year: int, month: int) -> List[date]:
    first = date(year, month, 1)
    last  = date(year, month, _cal.monthrange(year, month)[1])
    start = first - timedelta(days=first.weekday())
    end   = last  + timedelta(days=6 - last.weekday())
    days: List[date] = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)
    return days


def _iso_weeks(days: List[date]) -> List[List[date]]:
    seen: Dict[tuple, List[date]] = {}
    for d in days:
        key = d.isocalendar()[:2]
        seen.setdefault(key, []).append(d)
    return list(seen.values())


def _set(ws, row: int, col: int, value, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell


def build_calendar_filename(req: CalendarExportRequest) -> str:
    nombre_slug = req.branch_nombre.lower().replace(" ", "_")[:20]
    return f"calendario_{req.codigo_area}_{nombre_slug}_{req.year}{req.month:02d}.xlsx"


def export_calendar_to_xlsx(req: CalendarExportRequest) -> bytes:
    all_days  = _dias_iso_del_mes(req.year, req.month)
    weeks     = _iso_weeks(all_days)

    # Índices rápidos
    assign_index: Dict[tuple, str] = {}   # (slot, date_iso) -> "HH:MM-HH:MM"
    hours_index:  Dict[tuple, float] = {} # (slot, date_iso) -> duracion_h
    for a in req.assignments:
        assign_index[(a.slot, a.date)] = f"{a.inicio}–{a.fin}"
        h, m1 = map(int, a.inicio.split(":"))
        h2, m2 = map(int, a.fin.split(":"))
        duration = (h2 * 60 + m2 - h * 60 - m1) / 60
        hours_index[(a.slot, a.date)] = max(0.0, duration)

    n_workers = len(req.workers)
    slot_nombre = {w.slot: w.nombre for w in req.workers}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{req.codigo_area}_{req.year}{req.month:02d}"

    # Anchos de columna
    ws.column_dimensions["A"].width = 22  # nombre trabajador
    for col in range(2, 10):              # 7 días + horas
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(9)].width = 9

    row = 1

    # Fila título
    title_text = (
        f"CALENDARIO DE TURNOS  —  {req.branch_nombre.upper()}  —  "
        f"{MONTH_NAMES_ES[req.month].upper()} {req.year}"
    )
    _set(ws, row, 1, title_text, font=_FONT_TITLE, fill=_FILL_TITLE, align=_ALIGN_LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.row_dimensions[row].height = 22
    row += 1

    row += 1  # espacio en blanco

    for week_days in weeks:
        week_dates = [d.isoformat() for d in week_days]
        iso_week   = week_days[0].isocalendar()[1]
        first_d    = week_days[0]
        last_d     = week_days[-1]

        def fmt_date_range(d1: date, d2: date) -> str:
            if d1.month == d2.month:
                return f"{d1.day:02d} – {d2.day:02d} {MONTH_NAMES_ES[d2.month][:3]}"
            return (
                f"{d1.day:02d} {MONTH_NAMES_ES[d1.month][:3]} – "
                f"{d2.day:02d} {MONTH_NAMES_ES[d2.month][:3]}"
            )

        # Cabecera de semana
        week_label = f"  Sem {iso_week}   {fmt_date_range(first_d, last_d)}"
        _set(ws, row, 1, week_label, font=_FONT_WEEK, fill=_FILL_WEEK, align=_ALIGN_LEFT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 18
        row += 1

        # Cabecera de días
        _set(ws, row, 1, "Trabajador", font=_FONT_HEADER, fill=_FILL_HEADER,
             align=_ALIGN_CENTER, border=_BORDER_THIN)
        for ci, d in enumerate(week_days):
            is_wknd = d.weekday() >= 5
            label   = f"{WEEKDAY_ABBR[d.weekday()]} {d.day:02d}"
            fill    = _FILL_WEEKEND if is_wknd else _FILL_HEADER
            _set(ws, row, ci + 2, label, font=_FONT_HEADER, fill=fill,
                 align=_ALIGN_CENTER, border=_BORDER_THIN)
        _set(ws, row, 9, "Hrs Sem", font=_FONT_HEADER, fill=_FILL_HEADER,
             align=_ALIGN_CENTER, border=_BORDER_THIN)
        ws.row_dimensions[row].height = 16
        row += 1

        # Fila por trabajador
        for wi, slot in enumerate(range(1, n_workers + 1)):
            nombre = slot_nombre.get(slot, f"Trabajador {slot}")
            fill_row = _FILL_ALT if wi % 2 == 1 else _FILL_NONE
            _set(ws, row, 1, nombre, font=_FONT_NAME, fill=fill_row,
                 align=_ALIGN_LEFT, border=_BORDER_THIN)

            total_h = 0.0
            for ci, d in enumerate(week_days):
                date_str = d.isoformat()
                shift_str = assign_index.get((slot, date_str))
                h = hours_index.get((slot, date_str), 0.0)
                total_h += h
                if shift_str:
                    _set(ws, row, ci + 2, shift_str, font=_FONT_SHIFT, fill=fill_row,
                         align=_ALIGN_CENTER, border=_BORDER_THIN)
                else:
                    _set(ws, row, ci + 2, "libre", font=_FONT_LIBRE, fill=fill_row,
                         align=_ALIGN_CENTER, border=_BORDER_THIN)

            hrs_label = f"{total_h:.1f}h" if total_h > 0 else "—"
            _set(ws, row, 9, hrs_label, font=_FONT_HOURS, fill=fill_row,
                 align=_ALIGN_RIGHT, border=_BORDER_THIN)
            ws.row_dimensions[row].height = 15
            row += 1

        row += 1  # separador entre semanas

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
