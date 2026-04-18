"""Tests para excel_exporter.py — cubre los casos del DoD de Task 3."""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.schemas import Proposal, ShiftCatalog, Worker
from app.services.excel_exporter import (
    build_filename,
    export_proposal_to_xlsx,
    rut_without_dv,
)
from app.services.proposal_fetcher import ExportDataset, ResolvedAssignment


# ─── Helpers ──────────────────────────────────────────────────────────────────

def make_shift(shift_id: str, inicio: str = "09:00", fin: str = "19:00") -> ShiftCatalog:
    return ShiftCatalog.model_validate({
        "$id": shift_id, "nombre_display": shift_id,
        "hora_inicio": inicio, "hora_fin": fin,
        "duracion_minutos": 600, "descuenta_colacion": True, "categoria": "principal",
    })


def make_worker(worker_id: str, rut: str, nombre: str = "TRABAJADOR TEST") -> Worker:
    return Worker.model_validate({
        "$id": worker_id, "rut": rut, "nombre_completo": nombre,
        "branch_id": "b1", "activo": True,
    })


def make_dataset(
    year: int,
    month: int,
    resolved: list[ResolvedAssignment],
    workers: list[Worker],
    shifts: list[ShiftCatalog],
    codigo_area: str = "1200",
    branch_nombre: str = "NISSAN IRARRAZAVAL",
) -> ExportDataset:
    return ExportDataset(
        proposal=Proposal.model_validate({
            "$id": "p1", "branch_id": "b1", "anio": year, "mes": month,
            "modo": "ilp", "score": 90.0, "factible": True, "asignaciones": [],
            "dotacion_sugerida": 2, "parametros": {}, "estado": "seleccionada",
            "creada_por": "admin1",
        }),
        resolved_assignments=resolved,
        workers_by_id={w.id: w for w in workers},
        shifts_by_id={s.id: s for s in shifts},
        codigo_area=codigo_area,
        branch_nombre=branch_nombre,
        year=year,
        month=month,
    )


def load_sheet(xlsx_bytes: bytes):
    wb = load_workbook(BytesIO(xlsx_bytes))
    return wb.active


# ─── rut_without_dv ───────────────────────────────────────────────────────────

def test_rut_without_dv_numerico():
    assert rut_without_dv("17286931-9") == "17286931"


def test_rut_without_dv_k():
    assert rut_without_dv("12345678-K") == "12345678"


# ─── Headers y nombre de hoja ─────────────────────────────────────────────────

def test_sheet_name_formato():
    ds = make_dataset(2026, 5, [], [], [])
    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.title == "1200_202605"


def test_headers_mayo_31_dias():
    ds = make_dataset(2026, 5, [], [], [])
    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws["A1"].value == "RUT"
    assert ws.cell(row=1, column=2).value == "DIA1"
    assert ws.cell(row=1, column=32).value == "DIA31"
    assert ws.cell(row=1, column=33).value is None   # no hay columna 33


def test_headers_abril_30_dias():
    ds = make_dataset(2026, 4, [], [], [])
    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.cell(row=1, column=31).value == "DIA30"
    assert ws.cell(row=1, column=32).value is None


def test_headers_febrero_28_dias():
    ds = make_dataset(2026, 2, [], [], [])
    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.cell(row=1, column=29).value == "DIA28"
    assert ws.cell(row=1, column=30).value is None


def test_headers_febrero_29_dias_bisiesto():
    ds = make_dataset(2024, 2, [], [], [])
    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.cell(row=1, column=30).value == "DIA29"
    assert ws.cell(row=1, column=31).value is None


# ─── Contenido de celdas ──────────────────────────────────────────────────────

def test_turno_escrito_como_texto():
    shift = make_shift("S_09_19", "09:00", "19:00")
    worker = make_worker("w1", "17286931-9")
    resolved = [ResolvedAssignment(date="2026-05-05", shift_id="S_09_19", worker_id="w1")]
    ds = make_dataset(2026, 5, resolved, [worker], [shift])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    # DIA5 → columna 6
    assert ws.cell(row=2, column=6).value == "09:00 a 19:00"


def test_celda_vacia_cuando_no_hay_turno():
    shift = make_shift("S_09_19")
    worker = make_worker("w1", "17286931-9")
    resolved = [ResolvedAssignment(date="2026-05-01", shift_id="S_09_19", worker_id="w1")]
    ds = make_dataset(2026, 5, resolved, [worker], [shift])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    # DIA2 → columna 3, no tiene turno
    assert ws.cell(row=2, column=3).value is None


def test_rut_sin_dv_en_columna_a():
    shift = make_shift("S_09_19")
    worker = make_worker("w1", "17286931-9")
    resolved = [ResolvedAssignment(date="2026-05-01", shift_id="S_09_19", worker_id="w1")]
    ds = make_dataset(2026, 5, resolved, [worker], [shift])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.cell(row=2, column=1).value == "17286931"


def test_rut_con_dv_k():
    shift = make_shift("S_09_19")
    worker = make_worker("w1", "12345678-K")
    resolved = [ResolvedAssignment(date="2026-05-01", shift_id="S_09_19", worker_id="w1")]
    ds = make_dataset(2026, 5, resolved, [worker], [shift])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    assert ws.cell(row=2, column=1).value == "12345678"


def test_trabajador_sin_turno_no_aparece():
    shift = make_shift("S_09_19")
    worker_con = make_worker("w1", "11111111-1")
    worker_sin = make_worker("w2", "22222222-2")
    resolved = [ResolvedAssignment(date="2026-05-01", shift_id="S_09_19", worker_id="w1")]
    # w2 no tiene ningún resolved_assignment → no debe aparecer en el Excel
    ds = make_dataset(2026, 5, resolved, [worker_con, worker_sin], [shift])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    ruts = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "22222222" not in ruts
    assert "11111111" in ruts


def test_multiples_trabajadores_y_dias():
    shift_a = make_shift("S_09_19", "09:00", "19:00")
    shift_b = make_shift("S_10_18", "10:00", "18:00")
    w1 = make_worker("w1", "11111111-1")
    w2 = make_worker("w2", "22222222-2")
    resolved = [
        ResolvedAssignment(date="2026-05-03", shift_id="S_09_19", worker_id="w1"),
        ResolvedAssignment(date="2026-05-04", shift_id="S_10_18", worker_id="w1"),
        ResolvedAssignment(date="2026-05-03", shift_id="S_10_18", worker_id="w2"),
    ]
    ds = make_dataset(2026, 5, resolved, [w1, w2], [shift_a, shift_b])

    ws = load_sheet(export_proposal_to_xlsx(ds))
    # Verificar que hay exactamente 2 filas de datos (fila 2 y 3)
    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    assert len(data_rows) == 2


# ─── build_filename ───────────────────────────────────────────────────────────

def test_filename_formato():
    ds = make_dataset(2026, 5, [], [], [], codigo_area="1200", branch_nombre="NISSAN IRARRÁZAVAL")
    assert build_filename(ds) == "turnos_1200_nissan-irarrazaval_202605.xlsx"


def test_filename_sin_tildes():
    ds = make_dataset(2026, 4, [], [], [], codigo_area="9999", branch_nombre="Sucursal Ñuñoa")
    assert build_filename(ds) == "turnos_9999_sucursal-nunoa_202604.xlsx"
