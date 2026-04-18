"""
Test E2E de exportación (Task 9 — Spec 007).

Flujo:
  POST /optimize → propuesta real en memoria
  → ExportDataset construido desde esa propuesta (sin Appwrite)
  → POST /export (fetch_export_dataset y audit_log mockeados)
  → validación celda por celda del xlsx
"""
from __future__ import annotations

import re
from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.models.schemas import Proposal, ShiftCatalog, Worker
from app.services.proposal_fetcher import ExportDataset, ExportError, ResolvedAssignment

# ─── Datos de referencia ──────────────────────────────────────────────────────

_OPTIMIZE_PAYLOAD = {
    "branch": {
        "id": "b1",
        "codigo_area": "1200",
        "nombre": "NISSAN IRARRAZAVAL",
        "tipo_franja": "autopark",
    },
    "month": {"year": 2026, "month": 5},
    "workers": [
        {"rut": "17286931-9", "nombre": "ABARZUA VARGAS ANDREA", "constraints": []},
        {"rut": "12345678-9", "nombre": "GONZALEZ PEREZ CARLOS", "constraints": []},
        {"rut": "98765432-K", "nombre": "LOPEZ RAMIREZ MARIA",   "constraints": []},
    ],
    "holidays": ["2026-05-01", "2026-05-21"],
    "shift_catalog": [
        {"id": "S_09_19", "inicio": "09:00", "fin": "19:00", "duracion_minutos": 600, "descuenta_colacion": True},
        {"id": "S_10_20", "inicio": "10:00", "fin": "20:00", "duracion_minutos": 600, "descuenta_colacion": True},
    ],
    "franja_por_dia": {
        "lunes":     {"apertura": "09:00", "cierre": "19:00"},
        "martes":    {"apertura": "09:00", "cierre": "19:00"},
        "miercoles": {"apertura": "09:00", "cierre": "19:00"},
        "jueves":    {"apertura": "09:00", "cierre": "19:00"},
        "viernes":   {"apertura": "09:00", "cierre": "19:00"},
        "sabado":    {"apertura": "09:00", "cierre": "14:00"},
        "domingo":   None,
    },
    "parametros": {
        "modo": "greedy",
        "num_propuestas": 1,
        "horas_semanales_max": 42,
        "cobertura_minima": 1,
    },
}

# Workers locales: id estable para el dataset
_WORKERS_BY_RUT: dict[str, Worker] = {
    w["rut"]: Worker.model_validate({
        "$id": f"w{i + 1}",
        "rut": w["rut"],
        "nombre_completo": w["nombre"],
        "branch_id": "b1",
        "activo": True,
    })
    for i, w in enumerate(_OPTIMIZE_PAYLOAD["workers"])
}

_SHIFTS_BY_ID: dict[str, ShiftCatalog] = {
    s["id"]: ShiftCatalog.model_validate({
        "$id": s["id"],
        "nombre_display": s["id"],
        "hora_inicio": s["inicio"],
        "hora_fin": s["fin"],
        "duracion_minutos": s["duracion_minutos"],
        "descuenta_colacion": s["descuenta_colacion"],
        "categoria": "principal",
    })
    for s in _OPTIMIZE_PAYLOAD["shift_catalog"]
}


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def optimize_proposal(client):
    """Propuesta real generada por POST /optimize (sin Appwrite)."""
    resp = client.post("/optimize", json=_OPTIMIZE_PAYLOAD)
    assert resp.status_code == 200, resp.text
    propuestas = resp.json()["propuestas"]
    assert propuestas, "El optimizer no generó ninguna propuesta"
    return propuestas[0]


@pytest.fixture(scope="module")
def export_dataset(optimize_proposal) -> ExportDataset:
    """
    ExportDataset construido a partir de la propuesta del optimizador.
    Usa workers y turnos locales; no requiere conexión a Appwrite.
    """
    asignaciones = optimize_proposal["asignaciones"]

    resolved = [
        ResolvedAssignment(
            date=a["date"],
            shift_id=a["shift_id"],
            worker_id=_WORKERS_BY_RUT[a["worker_rut"]].id,
        )
        for a in asignaciones
        if a["worker_rut"] in _WORKERS_BY_RUT
    ]

    proposal = Proposal.model_validate({
        "$id": "prop_e2e",
        "branch_id": "b1",
        "anio": 2026,
        "mes": 5,
        "modo": "greedy",
        "score": optimize_proposal["score"],
        "factible": True,
        "asignaciones": [
            {"slot": i + 1, "date": a["date"], "shift_id": a["shift_id"]}
            for i, a in enumerate(asignaciones)
        ],
        "dotacion_sugerida": len(_OPTIMIZE_PAYLOAD["workers"]),
        "parametros": {},
        "estado": "seleccionada",
        "creada_por": "test-admin",
    })

    return ExportDataset(
        proposal=proposal,
        resolved_assignments=resolved,
        workers_by_id={w.id: w for w in _WORKERS_BY_RUT.values()},
        shifts_by_id=_SHIFTS_BY_ID,
        codigo_area="1200",
        branch_nombre="NISSAN IRARRAZAVAL",
        year=2026,
        month=5,
    )


@pytest.fixture(scope="module")
def xlsx_response(client, export_dataset):
    """
    Llama POST /export una vez y retorna la Response completa.
    fetch_export_dataset y create_audit_log se mockean para no depender de Appwrite.
    """
    with (
        patch("app.api.routes.fetch_export_dataset", new=AsyncMock(return_value=export_dataset)),
        patch("app.services.appwrite_client.create_audit_log", new=AsyncMock()),
    ):
        resp = client.post("/export", json={"proposal_id": "prop_e2e"})
    assert resp.status_code == 200, resp.text
    return resp


@pytest.fixture(scope="module")
def xlsx_ws(xlsx_response):
    return load_workbook(BytesIO(xlsx_response.content)).active


# ─── Happy path ───────────────────────────────────────────────────────────────

class TestExportHappyPath:

    def test_status_200_y_content_type_xlsx(self, xlsx_response):
        assert xlsx_response.status_code == 200
        assert "spreadsheetml" in xlsx_response.headers["content-type"]

    def test_content_disposition_attachment_con_filename(self, xlsx_response):
        cd = xlsx_response.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd
        assert "1200" in cd

    def test_xlsx_valido_y_hoja_activa(self, xlsx_ws):
        assert xlsx_ws is not None

    def test_nombre_hoja_formato_area_yyyymm(self, xlsx_response):
        wb = load_workbook(BytesIO(xlsx_response.content))
        assert "1200_202605" in wb.sheetnames

    def test_header_a1_es_rut(self, xlsx_ws):
        assert xlsx_ws["A1"].value == "RUT"

    def test_header_dia1_en_b1(self, xlsx_ws):
        assert xlsx_ws.cell(row=1, column=2).value == "DIA1"

    def test_header_dia31_en_columna_32_mayo(self, xlsx_ws):
        # Mayo tiene 31 días → DIA31 en columna 32 (A=1, DIA1=2, ..., DIA31=32)
        assert xlsx_ws.cell(row=1, column=32).value == "DIA31"

    def test_no_hay_columna_33(self, xlsx_ws):
        assert xlsx_ws.cell(row=1, column=33).value is None

    def test_ruts_sin_guion_ni_dv(self, xlsx_ws):
        for row in range(2, xlsx_ws.max_row + 1):
            val = xlsx_ws.cell(row=row, column=1).value
            if val:
                assert "-" not in str(val), f"RUT con guión en fila {row}: {val!r}"

    def test_ruts_son_numericos(self, xlsx_ws):
        for row in range(2, xlsx_ws.max_row + 1):
            val = xlsx_ws.cell(row=row, column=1).value
            if val:
                assert str(val).isdigit(), f"RUT no numérico en fila {row}: {val!r}"

    def test_turnos_formato_hhmm_a_hhmm(self, xlsx_ws):
        patron = re.compile(r"^\d{2}:\d{2} a \d{2}:\d{2}$")
        for row in range(2, xlsx_ws.max_row + 1):
            for col in range(2, xlsx_ws.max_column + 1):
                val = xlsx_ws.cell(row=row, column=col).value
                if val is not None:
                    assert patron.match(str(val)), (
                        f"Formato incorrecto en ({row},{col}): {val!r}"
                    )

    def test_cada_fila_tiene_al_menos_un_turno(self, xlsx_ws):
        for row in range(2, xlsx_ws.max_row + 1):
            rut = xlsx_ws.cell(row=row, column=1).value
            if rut:
                tiene_turno = any(
                    xlsx_ws.cell(row=row, column=c).value is not None
                    for c in range(2, xlsx_ws.max_column + 1)
                )
                assert tiene_turno, f"Fila {row} (RUT {rut!r}) no tiene ningún turno"

    def test_hay_al_menos_una_fila_de_datos(self, xlsx_ws):
        filas_con_rut = sum(
            1 for row in range(2, xlsx_ws.max_row + 1)
            if xlsx_ws.cell(row=row, column=1).value
        )
        assert filas_con_rut >= 1


# ─── Error path ───────────────────────────────────────────────────────────────

class TestExportErrorPath:

    def test_422_cuando_propuesta_no_seleccionada(self, client):
        with patch(
            "app.api.routes.fetch_export_dataset",
            new=AsyncMock(side_effect=ExportError("propuesta no seleccionada")),
        ):
            resp = client.post("/export", json={"proposal_id": "cualquiera"})
        assert resp.status_code == 422

    def test_422_cuando_slots_sin_asignar(self, client):
        with patch(
            "app.api.routes.fetch_export_dataset",
            new=AsyncMock(side_effect=ExportError("La propuesta tiene 2 slot(s) sin asignar: [1, 2]")),
        ):
            resp = client.post("/export", json={"proposal_id": "cualquiera"})
        assert resp.status_code == 422

    def test_404_cuando_propuesta_no_existe(self, client):
        with patch(
            "app.api.routes.fetch_export_dataset",
            new=AsyncMock(side_effect=KeyError("no existe")),
        ):
            resp = client.post("/export", json={"proposal_id": "no_existe"})
        assert resp.status_code == 404
