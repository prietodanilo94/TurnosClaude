import asyncio
import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.schemas import (
    Assignment,
    Branch,
    Holiday,
    OverrideType,
    Proposal,
    ShiftCatalogV2,
    SlotOverride,
    Worker,
)
from app.services import appwrite_client as ac
from app.services.excel_exporter import EM_DASH, export_proposal_to_xlsx
from app.services.proposal_fetcher import (
    ExportDataset,
    ExportError,
    ResolvedAssignment,
    fetch_export_dataset,
)


def _proposal_doc(**overrides):
    base = {
        "$id": "prop_1",
        "branch_id": "branch_1",
        "anio": 2026,
        "mes": 5,
        "modo": "ilp",
        "score": 98.5,
        "factible": True,
        "asignaciones": json.dumps(
            [{"slot": 1, "date": "2026-05-04", "shift_id": "ape", "worker_rut": "1-9"}]
        ),
        "dotacion_sugerida": 3,
        "parametros": json.dumps({"modo": "ilp"}),
        "estado": "seleccionada",
        "creada_por": "user_1",
        "metrics": json.dumps({"cobertura_peak_pct": 100}),
    }
    base.update(overrides)
    return base


def _worker_doc(**overrides):
    base = {
        "$id": "worker_1",
        "rut": "17286931-9",
        "nombre_completo": "Andrea Prueba",
        "branch_id": "branch_1",
        "activo": True,
    }
    base.update(overrides)
    return Worker.model_validate(base)


def _override_doc(**overrides):
    base = {
        "$id": "override_1",
        "proposal_id": "prop_1",
        "fecha": "2026-05-01",
        "slot_numero": 1,
        "tipo": "cambiar_turno",
        "notas": "Ajuste manual",
        "creado_por": "user_1",
    }
    base.update(overrides)
    return SlotOverride.model_validate(base)


def _dataset_for_export(
    *,
    resolved_assignments: list[ResolvedAssignment] | None = None,
    slot_overrides: list[SlotOverride] | None = None,
    holidays: set[str] | None = None,
    closed_dates: set[str] | None = None,
) -> ExportDataset:
    proposal = Proposal.model_validate(_proposal_doc())
    worker = _worker_doc()
    return ExportDataset(
        proposal=proposal,
        resolved_assignments=resolved_assignments
        or [
            ResolvedAssignment(
                slot_numero=1,
                date="2026-05-01",
                shift_id="ape",
                shift_label="Apertura corta",
                worker_id="worker_1",
                hora_inicio="11:00",
                hora_fin="20:00",
                horas_laborales=8.0,
            )
        ],
        workers_by_id={"worker_1": worker},
        worker_slot_by_id={"worker_1": 1},
        codigo_area="107",
        branch_nombre="Movicenter",
        branch_clasificacion="Mall",
        area_negocio_label="Ventas",
        year=2026,
        month=5,
        holidays=holidays or set(),
        closed_dates=closed_dates or set(),
        slot_overrides=slot_overrides or [],
    )


def test_proposal_parses_json_string_fields():
    proposal = Proposal.model_validate(_proposal_doc())

    assert proposal.asignaciones[0].slot == 1
    assert proposal.asignaciones[0].date == "2026-05-04"
    assert proposal.parametros == {"modo": "ilp"}
    assert proposal.metrics == {"cobertura_peak_pct": 100}


def test_fetch_export_dataset_resolves_shift_catalog_v2_hours(monkeypatch):
    proposal = Proposal.model_validate(_proposal_doc())
    branch = Branch.model_validate(
        {
            "$id": "branch_1",
            "codigo_area": "107",
            "nombre": "Movicenter",
            "tipo_franja": "movicenter",
            "clasificacion": "Mall",
            "activa": True,
            "creada_desde_excel": True,
        }
    )
    worker = _worker_doc(area_negocio="ventas", rotation_group="V_M7")
    shift = ShiftCatalogV2.model_validate(
        {
            "$id": "ape",
            "nombre_display": "Apertura corta",
            "rotation_group": "V_M7",
            "nombre_turno": "apertura",
            "horario_por_dia": json.dumps(
                {
                    "lunes": {"inicio": "10:00", "fin": "19:00"},
                    "domingo": {"inicio": "11:00", "fin": "20:00"},
                }
            ),
            "descuenta_colacion": True,
            "dias_aplicables": ["lunes", "domingo"],
        }
    )
    assignment = Assignment.model_validate(
        {
            "$id": "asg_1",
            "proposal_id": "prop_1",
            "slot_numero": 1,
            "worker_id": "worker_1",
        }
    )
    holiday = Holiday.model_validate(
        {
            "$id": "holiday_1",
            "fecha": "2026-05-21",
            "nombre": "Glorias Navales",
            "tipo": "irrenunciable",
            "anio": 2026,
        }
    )
    override = _override_doc(fecha="2026-05-04")

    async def fake_get_proposal(_proposal_id: str):
        return proposal

    async def fake_list_assignments(_proposal_id: str):
        return [assignment]

    async def fake_get_branch(_branch_id: str):
        return branch

    async def fake_get_shift_catalog_v2():
        return [shift]

    async def fake_list_workers_by_branch(_branch_id: str):
        return [worker]

    async def fake_list_holidays_by_year(_year: int):
        return [holiday]

    async def fake_list_slot_overrides_by_proposal(_proposal_id: str):
        return [override]

    async def fake_list_workers(_worker_ids: list[str]):
        return [worker]

    monkeypatch.setattr(ac, "get_proposal", fake_get_proposal)
    monkeypatch.setattr(ac, "list_assignments_by_proposal", fake_list_assignments)
    monkeypatch.setattr(ac, "get_branch", fake_get_branch)
    monkeypatch.setattr(ac, "get_shift_catalog_v2", fake_get_shift_catalog_v2)
    monkeypatch.setattr(ac, "list_workers_by_branch", fake_list_workers_by_branch)
    monkeypatch.setattr(ac, "list_holidays_by_year", fake_list_holidays_by_year)
    monkeypatch.setattr(ac, "list_slot_overrides_by_proposal", fake_list_slot_overrides_by_proposal)
    monkeypatch.setattr(ac, "list_workers_by_ids", fake_list_workers)

    dataset = asyncio.run(fetch_export_dataset("prop_1"))

    assert dataset.codigo_area == "107"
    assert dataset.branch_nombre == "Movicenter"
    assert dataset.branch_clasificacion == "Mall"
    assert dataset.area_negocio_label == "Ventas"
    assert dataset.worker_slot_by_id == {"worker_1": 1}
    assert dataset.holidays == {"2026-05-21"}
    assert dataset.slot_overrides == [override]
    assert dataset.resolved_assignments == [
        ResolvedAssignment(
            slot_numero=1,
            date="2026-05-04",
            shift_id="ape",
            shift_label="Apertura corta",
            worker_id="worker_1",
            hora_inicio="10:00",
            hora_fin="19:00",
            horas_laborales=8.0,
            override_applied=True,
            override_type="cambiar_turno",
            override_note="Ajuste manual",
        )
    ]


def test_fetch_export_dataset_rejects_unassigned_slots(monkeypatch):
    proposal = Proposal.model_validate(_proposal_doc())
    branch = Branch.model_validate(
        {
            "$id": "branch_1",
            "codigo_area": "107",
            "nombre": "Movicenter",
            "tipo_franja": "movicenter",
            "clasificacion": "Mall",
            "activa": True,
            "creada_desde_excel": True,
        }
    )
    shift = ShiftCatalogV2.model_validate(
        {
            "$id": "ape",
            "nombre_display": "Apertura corta",
            "rotation_group": "V_M7",
            "nombre_turno": "apertura",
            "horario_por_dia": {"lunes": {"inicio": "10:00", "fin": "19:00"}},
            "descuenta_colacion": True,
            "dias_aplicables": ["lunes"],
        }
    )
    assignment = Assignment.model_validate(
        {
            "$id": "asg_1",
            "proposal_id": "prop_1",
            "slot_numero": 1,
            "worker_id": None,
        }
    )

    async def fake_get_proposal(_proposal_id: str):
        return proposal

    async def fake_list_assignments(_proposal_id: str):
        return [assignment]

    async def fake_get_branch(_branch_id: str):
        return branch

    async def fake_get_shift_catalog_v2():
        return [shift]

    async def fake_list_workers_by_branch(_branch_id: str):
        return []

    async def fake_list_holidays_by_year(_year: int):
        return []

    async def fake_list_slot_overrides_by_proposal(_proposal_id: str):
        return []

    async def fake_list_workers(_worker_ids: list[str]):
        return []

    monkeypatch.setattr(ac, "get_proposal", fake_get_proposal)
    monkeypatch.setattr(ac, "list_assignments_by_proposal", fake_list_assignments)
    monkeypatch.setattr(ac, "get_branch", fake_get_branch)
    monkeypatch.setattr(ac, "get_shift_catalog_v2", fake_get_shift_catalog_v2)
    monkeypatch.setattr(ac, "list_workers_by_branch", fake_list_workers_by_branch)
    monkeypatch.setattr(ac, "list_holidays_by_year", fake_list_holidays_by_year)
    monkeypatch.setattr(ac, "list_slot_overrides_by_proposal", fake_list_slot_overrides_by_proposal)
    monkeypatch.setattr(ac, "list_workers_by_ids", fake_list_workers)

    with pytest.raises(ExportError, match="slot\\(s\\) sin asignar"):
        asyncio.run(fetch_export_dataset("prop_1"))


def test_export_proposal_to_xlsx_marks_override_and_uses_labor_hours():
    dataset = _dataset_for_export(
        resolved_assignments=[
            ResolvedAssignment(
                slot_numero=1,
                date="2026-05-01",
                shift_id="ape",
                shift_label="Apertura corta",
                worker_id="worker_1",
                hora_inicio="11:00",
                hora_fin="20:00",
                horas_laborales=8.0,
                override_applied=True,
                override_type=OverrideType.cambiar_turno.value,
                override_note="Ajuste manual",
            )
        ],
        slot_overrides=[_override_doc()],
    )

    content = export_proposal_to_xlsx(dataset)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert ws["A1"].value == "Turno Mensual - Movicenter - Mayo 2026"
    assert ws["A2"].value == "Clasificacion: Mall"
    assert ws["A3"].value == "Area de Negocio: Ventas"
    assert str(ws["A4"].value).startswith("Generado: ")
    assert ws["A6"].value == "RUT"
    assert ws["C6"].value == "Turno Base"
    assert ws["A7"].value == "17286931"
    assert ws["C7"].value == "Apertura corta"
    assert ws["D7"].value == "11:00-20:00 *"
    assert ws.cell(row=7, column=35).value == 8
    assert ws["A10"].value == "Notas overrides"
    assert "2026-05-01 slot 1: cambiar_turno" in ws["A11"].value


def test_export_proposal_to_xlsx_shows_human_readable_times_not_shift_ids():
    dataset = _dataset_for_export()

    content = export_proposal_to_xlsx(dataset)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert ws["D7"].value == "11:00-20:00"
    assert "ape" not in str(ws["D7"].value)


def test_export_proposal_to_xlsx_keeps_sunday_column_for_mall_7d():
    dataset = _dataset_for_export()

    content = export_proposal_to_xlsx(dataset)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert ws.cell(row=6, column=34).value == 31


def test_export_proposal_to_xlsx_marks_closed_day_with_dash():
    dataset = _dataset_for_export(closed_dates={"2026-05-31"})

    content = export_proposal_to_xlsx(dataset)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert ws.cell(row=7, column=34).value == EM_DASH


def test_export_proposal_to_xlsx_marks_holiday():
    dataset = _dataset_for_export(holidays={"2026-05-21"})

    content = export_proposal_to_xlsx(dataset)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert ws.cell(row=7, column=24).value == "FERIADO"
