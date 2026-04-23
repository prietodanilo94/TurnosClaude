"""
Genera el archivo .xlsx del turnero a partir de un ExportDataset.
"""
from __future__ import annotations

import re
import unicodedata
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.proposal_fetcher import ExportDataset, ResolvedAssignment

EM_DASH = "\u2014"
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HOLIDAY_FILL = PatternFill("solid", fgColor="FDE68A")
FREE_FILL = PatternFill("solid", fgColor="E5E7EB")
CLOSED_FILL = PatternFill("solid", fgColor="9CA3AF")
WORK_FILL = PatternFill("solid", fgColor="2E7D32")
WORK_FONT = Font(color="FFFFFF")
MONTH_NAMES_ES = (
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)


def rut_without_dv(rut: str) -> str:
    """'17286931-9' -> '17286931', '12345678-K' -> '12345678'."""
    return rut.split("-")[0]


def _slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    return normalized.strip("-")


def build_filename(dataset: ExportDataset) -> str:
    slug = _slugify(dataset.branch_nombre)
    return f"turnos_{dataset.codigo_area}_{slug}_{dataset.year}{dataset.month:02d}.xlsx"


def _month_title(dataset: ExportDataset) -> str:
    month_label = MONTH_NAMES_ES[dataset.month]
    return f"Turno Mensual - {dataset.branch_nombre} - {month_label} {dataset.year}"


def _group_assignments(
    assignments: list[ResolvedAssignment],
) -> dict[str, dict[int, ResolvedAssignment]]:
    grouped: dict[str, dict[int, ResolvedAssignment]] = defaultdict(dict)
    for assignment in assignments:
        day = int(assignment.date[-2:])
        grouped[assignment.worker_id][day] = assignment
    return grouped


def _turno_base(assignments: dict[int, ResolvedAssignment]) -> str:
    labels = [assignment.shift_label for assignment in assignments.values()]
    if not labels:
        return ""
    return Counter(labels).most_common(1)[0][0]


def _override_notes(dataset: ExportDataset) -> list[str]:
    notes: list[str] = []
    ordered = sorted(dataset.slot_overrides, key=lambda item: (item.fecha, item.slot_numero or 0))
    for override in ordered:
        note = f"* {override.fecha} slot {override.slot_numero}: {override.tipo.value}"
        if override.notas:
            note += f" - {override.notas}"
        notes.append(note)
    return notes


def export_proposal_to_xlsx(dataset: ExportDataset) -> bytes:
    days_in_month = monthrange(dataset.year, dataset.month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{dataset.codigo_area}_{dataset.year}{dataset.month:02d}"

    ws["A1"] = _month_title(dataset)
    ws["A2"] = f"Clasificacion: {dataset.branch_clasificacion or 'No definida'}"
    ws["A3"] = f"Area de Negocio: {dataset.area_negocio_label}"
    ws["A4"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    header_row = 6
    ws.cell(row=header_row, column=1, value="RUT")
    ws.cell(row=header_row, column=2, value="Nombre")
    ws.cell(row=header_row, column=3, value="Turno Base")
    for day in range(1, days_in_month + 1):
        ws.cell(row=header_row, column=day + 3, value=day)
    ws.cell(row=header_row, column=days_in_month + 4, value="Horas Mes")

    for col in range(1, days_in_month + 5):
        ws.cell(row=header_row, column=col).fill = HEADER_FILL
        ws.cell(row=header_row, column=col).font = HEADER_FONT

    by_worker = _group_assignments(dataset.resolved_assignments)
    row = header_row + 1

    for worker_id in dataset.worker_slot_by_id:
        assignments_by_day = by_worker.get(worker_id, {})
        worker = dataset.workers_by_id.get(worker_id)
        if worker is None:
            continue

        ws.cell(row=row, column=1, value=rut_without_dv(worker.rut))
        ws.cell(row=row, column=2, value=worker.nombre_completo)
        ws.cell(row=row, column=3, value=_turno_base(assignments_by_day))

        total_hours = 0.0
        worker_slot = dataset.worker_slot_by_id.get(worker_id)

        for day in range(1, days_in_month + 1):
            iso_date = f"{dataset.year:04d}-{dataset.month:02d}-{day:02d}"
            cell = ws.cell(row=row, column=day + 3)
            assignment = assignments_by_day.get(day)

            if assignment:
                cell.value = f"{assignment.hora_inicio}-{assignment.hora_fin}"
                if assignment.override_applied:
                    cell.value += " *"
                cell.fill = WORK_FILL
                cell.font = WORK_FONT
                total_hours += assignment.horas_laborales
                continue

            protected_or_freed = any(
                override.slot_numero == worker_slot and override.fecha[:10] == iso_date
                for override in dataset.slot_overrides
            )

            if iso_date in dataset.holidays:
                cell.value = "FERIADO"
                cell.fill = HOLIDAY_FILL
            elif iso_date in dataset.closed_dates:
                cell.value = EM_DASH
                cell.fill = CLOSED_FILL
            else:
                cell.value = "LIBRE *" if protected_or_freed else "LIBRE"
                cell.fill = FREE_FILL

        ws.cell(row=row, column=days_in_month + 4, value=round(total_hours, 2))
        row += 1

    notes = _override_notes(dataset)
    if notes:
        start_row = row + 2
        ws.cell(row=start_row, column=1, value="Notas overrides")
        for index, note in enumerate(notes, start=1):
            ws.cell(row=start_row + index, column=1, value=note)

    ws.freeze_panes = "D7"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
